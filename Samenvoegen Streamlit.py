import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import copy
import io

st.title("Excel Samenvoeger (Layout + Logo behouden)")

uploaded_files = st.file_uploader(
    "Upload één of meerdere Excel-bestanden (.xlsx)", 
    type="xlsx", 
    accept_multiple_files=True
)

if uploaded_files:
    if st.button("Bestanden samenvoegen"):
        # Laad het eerste bestand als basis
        base_file = uploaded_files[0]
        base_wb = load_workbook(filename=io.BytesIO(base_file.read()))
        base_ws = base_wb.active

        # Voeg de rest van de bestanden toe
        for file in uploaded_files[1:]:
            file.seek(0)
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            # Vind de eerste lege rij in het basisbestand
            first_empty_row = base_ws.max_row + 1

            # Kopieer cellen + stijlen
            for i, row in enumerate(ws.iter_rows(), start=first_empty_row):
                for j, cell in enumerate(row, start=1):
                    new_cell = base_ws.cell(row=i, column=j, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)

            # Kopieer afbeeldingen (logo’s)
            for img in ws._images:
                new_img = Image(img.ref) if hasattr(img, 'ref') else Image(img._data())
                new_img.anchor = img.anchor
                base_ws.add_image(new_img)

        # Sla het samengevoegde bestand op in memory
        output = io.BytesIO()
        base_wb.save(output)
        output.seek(0)

        st.success("Bestanden succesvol samengevoegd (inclusief logo’s)!")
        st.download_button(
            label="Download samengevoegd bestand",
            data=output,
            file_name="samengevoegd.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
